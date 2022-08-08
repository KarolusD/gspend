from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.chart.axis import DateAxis
from openpyxl.chart.text import RichText
from openpyxl.chart.label import DataLabelList
from openpyxl.chart import PieChart, LineChart, Reference
from openpyxl.drawing.text import (
    Paragraph,
    ParagraphProperties,
    CharacterProperties,
    Font as DrawFont,
)

# built-in libraries
import csv
from datetime import datetime
from typing import Any, TypedDict

WHERE_TRANSACTIONS_START = "#Data operacji"
WHERE_TRANSACTIONS_END = "___End of file___"

class Transaction(TypedDict):
    date: datetime
    category: str
    amount: float


class Excel(TypedDict):
    workbook: Workbook
    sheets: dict[str, Any]


def generate_excel():
    try:
        path_to_CSV = input("Provide CSV file path:\n")

        print("\nReading CSV transactions...")
        all_transactions = read_csv_transactions(path_to_CSV)

        print("Creating basic excel structure...")
        excel = create_excel()

        print("Writing transactions into excel...")
        write_transactions_into_excel(excel["sheets"]["spendings"], all_transactions)

        print("Creating transactions line chart...")
        create_line_chart_for_transactions_in_time(
            excel["sheets"]["spendings"], len(all_transactions)
        )

        print("Grouping expenses by category...")
        expenses_by_category = group_transactions_by_categories(all_transactions)

        print("Writing expenses by category into excel...")
        write_expenses_by_category_into_excel(
            excel["sheets"]["expenses"], expenses_by_category
        )

        print("Creating expenses pie chart...\n")
        create_pie_chart_for_expenses_by_category(
            excel["sheets"]["expenses"], len(expenses_by_category)
        )

        path_to_excel = input("Where to save excel? (./gspend.xlsx)\n") or "./gspend.xlsx"

        excel["workbook"].save(path_to_excel)
        print("\nExcel file was successfully generated 🚀")
    except Exception as exception:
        print(exception)
        print("Something went wrong while generating excel file 😢")


generate_excel()


def create_excel() -> Excel:
    wb = Workbook()

    ws_spendings = wb.create_sheet("Spendings", 0)
    ws_expenses = wb.create_sheet("Expenses", 1)

    ws_spendings.cell(row=1, column=1, value="Date")
    ws_spendings.cell(row=1, column=2, value="Category")
    ws_spendings.cell(row=1, column=3, value="Amount")

    ws_expenses.cell(row=1, column=1, value="Category")
    ws_expenses.cell(row=1, column=2, value="Amount")

    return {
        "workbook": wb,
        "sheets": {"spendings": ws_spendings, "expenses": ws_expenses},
    }


# NOTE: other banks may have different CSV format
def read_csv_transactions(file_path: str) -> list[Transaction]:
    # encoding cp1250 is needed to read Polish chars outside UTF-8 range
    with open(file_path, "r", encoding="cp1250", errors="ignore") as file:
        # CSV delimiter in this case is ; often times CSV uses ,
        csv_reader = csv.reader(file, delimiter=";")
        start = False
        all_transactions = []

        for line in csv_reader:
            if line and line[0]:
                if  line[0] == WHERE_TRANSACTIONS_END:
                    start = False

                if start:

                    date = datetime.strptime(line[0], "%Y-%m-%d").strftime("%m/%d/%Y")
                    category = line[3]
                    amount = float(
                        line[4].replace(" PLN", "").replace(",", ".").replace(" ", "")
                    )

                    transaction = {"date": date, "category": category, "amount": amount}
                    all_transactions.append(transaction)

                # Check whether transactions are in the file and when they start
                if line[0] == WHERE_TRANSACTIONS_START:
                    start = True


        return all_transactions


def write_transactions_into_excel(sheet: Any, all_transactions: list[Transaction]):
    currency_format = "# ##0.00 [$zł-415]; [Red] -# ##0.00 [$zł-415]"
    for idx, transaction in enumerate(all_transactions):
        row = idx + 2  # rows start from 1 and first row is for header
        sheet.cell(row, column=1, value=transaction["date"])
        sheet.cell(row, column=2, value=transaction["category"])
        sheet.cell(row, column=3, value=transaction["amount"])

        sheet[f"C{row}"].number_format = currency_format

    transactions_len = len(all_transactions)
    last_row_B = f"B{transactions_len+2}"
    last_row_C = f"C{transactions_len+2}"

    sheet[last_row_B] = "Total"
    sheet[last_row_C] = f"=SUM(C2:C{transactions_len+1})"

    # Some styling (could be improved)
    font = Font(bold=True)
    sheet[last_row_B].font = font
    sheet[last_row_C].font = font
    sheet[last_row_C].number_format = currency_format

    # posibility to sort transactions by date
    sheet.auto_filter.ref = f"A1:C{transactions_len+1}"
    sheet.auto_filter.add_sort_condition(f"A2:A{transactions_len+1}")


def write_expenses_by_category_into_excel(
    sheet: Any, expenses_by_category: list[tuple]
):
    currency_format = "# ##0.00 [$zł-415]; [Red] -# ##0.00 [$zł-415]"
    categories_count = len(expenses_by_category)

    for idx, exp in enumerate(expenses_by_category):
        row = idx + 2

        sheet.cell(row, column=1, value=exp[0])
        sheet.cell(row, column=2, value=exp[1])

        sheet[f"B{row}"].number_format = currency_format

    font = Font(bold=True)

    sheet[f"A{categories_count+3}"] = "Total"
    sheet[f"A{categories_count+3}"].font = font

    sheet[f"B{categories_count+3}"] = f"=SUM(B1:B{categories_count})"
    sheet[f"B{categories_count+3}"].font = font


def group_transactions_by_categories(
    all_transactions: list[Transaction],
) -> list[tuple]:
    transactions_by_category = {}
    for transaction in all_transactions:
        category = transaction["category"]
        if category in transactions_by_category:
            transactions_by_category[category] += transaction["amount"]
        else:
            transactions_by_category[category] = transaction["amount"]

    expenses_by_category = {}
    for (key, value) in transactions_by_category.items():
        if value < 0:
            expenses_by_category[key] = value

    return sorted(expenses_by_category.items(), key=lambda x: x[1], reverse=True)


def create_pie_chart_for_expenses_by_category(sheet: Any, max_row: int):
    pie = PieChart()
    pie.title = "Expenses by category"
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = True  # show amount
    pie.dataLabels.showCatName = True

    pie.width = 60
    pie.height = 60

    labels = Reference(sheet, min_col=1, min_row=2, max_row=max_row)
    data = Reference(sheet, min_col=2, min_row=2, max_row=max_row)

    # https://stackoverflow.com/questions/56551838/with-python-openpyxl-how-do-you-change-the-font-size-in-a-chart-legend
    font = DrawFont(typeface="Verdana")
    size = 1600  # 16 point size
    cp = CharacterProperties(latin=font, sz=size, b=True)  # bold text
    pp = ParagraphProperties(defRPr=cp)
    rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
    pie.legend.textProperties = rtp

    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)

    sheet.add_chart(pie, "E1")


def create_line_chart_for_transactions_in_time(sheet: Any, max_row: int):
    chart = LineChart()
    chart.title = "Spendings in time"
    chart.style = 12
    chart.y_axis.title = "Amount"
    chart.y_axis.crossAx = 100
    chart.x_axis = DateAxis(crossAx=100)
    chart.x_axis.number_format = "d-m"
    chart.x_axis.majorTimeUnit = "days"
    chart.x_axis.title = "Date"
    chart.width = 60
    chart.height = 30

    data = Reference(sheet, min_col=3, min_row=2, max_row=max_row)
    dates = Reference(sheet, min_col=1, min_row=2, max_row=max_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(dates)
    chart.legend = None

    sheet.add_chart(chart, "E1")
