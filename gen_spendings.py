import csv
from openpyxl import Workbook
from openpyxl.descriptors import DateTime
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.chart import BarChart, Reference
from datetime import datetime
from typing import Any, TypedDict


class Transaction(TypedDict):
    date: datetime
    category: str
    amount: float


class Excel(TypedDict):
    workbook: Workbook
    sheets: dict[str, Any]


def create_excel() -> Excel:
    wb = Workbook()

    ws_spendings = wb.create_sheet("Spendings", 0)
    ws_summary = wb.create_sheet("Summary", 1)

    ws_spendings.cell(row=1, column=1, value="Date")
    ws_spendings.cell(row=1, column=2, value="Category")
    ws_spendings.cell(row=1, column=3, value="Price")

    return {
        "workbook": wb,
        "sheets": {"spendings": ws_spendings, "summary": ws_summary},
    }


# NOTE: other banks may have different CSV format
def read_csv_transactions(file_path: str) -> list[Transaction]:
    # encoding cp1250 is needed to read Polish chars outside UTF-8 range
    with open(file_path, "r", encoding="cp1250", errors="ignore") as file:
        # CSV delimiter in this case is ; often times CSV uses ,
        csv_reader = csv.reader(file, delimiter=";")
        start = False
        all_transactions = []

        for line in csv_reader:
            if line and line[0]:
                if start:
                    date = datetime.strptime(line[0], "%Y-%m-%d").strftime("%m/%d/%Y")
                    category = line[3]
                    amount = float(
                        line[4].replace(" PLN", "").replace(",", ".").replace(" ", "")
                    )

                    transaction = {"date": date, "category": category, "amount": amount}
                    all_transactions.append(transaction)

                # Check whether transactions are in the file and when they start
                if line[0] == "#Data operacji":
                    start = True

        return all_transactions


def write_transactions_into_excel(sheet: Any, all_transactions: list[Transaction]):
    currency_format = "# ##0.00 [$zł-415]; [Red] -# ##0.00 [$zł-415]"
    for idx, transaction in enumerate(all_transactions):
        row = idx + 2  # rows start from 1 and first row is for header
        sheet.cell(row, column=1, value=transaction["date"])
        sheet.cell(row, column=2, value=transaction["category"])
        sheet.cell(row, column=3, value=transaction["amount"])

        sheet[f"C{row}"].number_format = currency_format

    transactions_len = len(all_transactions)
    last_row_B = f"B{transactions_len+2}"
    last_row_C = f"C{transactions_len+2}"

    sheet[last_row_B] = "Total"
    sheet[last_row_C] = f"=SUM(C2:C{transactions_len+1})"

    # Some styling (could be improved)
    font = Font(bold=True)
    sheet[last_row_B].font = font
    sheet[last_row_C].font = font
    sheet[last_row_C].number_format = currency_format


def write_expenses_by_category_into_excel(
    sheet: Any, expenses_by_category: dict[str, Any]
):
    currency_format = "# ##0.00 [$zł-415]; [Red] -# ##0.00 [$zł-415]"
    categories_count = len(expenses_by_category)

    for idx, category in enumerate(expenses_by_category):
        row = idx + 1
        amount = expenses_by_category[category]

        sheet.cell(row, column=1, value=category)
        sheet.cell(row, column=2, value=amount)

        sheet[f"B{row}"].number_format = currency_format

    font = Font(bold=True)

    sheet[f"A{categories_count+1}"] = "Total"
    sheet[f"A{categories_count+1}"].font = font

    sheet[f"B{categories_count+1}"] = f"=SUM(B1:B{categories_count})"
    sheet[f"B{categories_count+1}"].font = font


def group_transactions_by_categories(
    all_transactions: list[Transaction],
) -> dict[str, float]:
    expenses_by_category = {}
    for transaction in all_transactions:
        category = transaction["category"]
        if category in expenses_by_category:
            expenses_by_category[category] += transaction["amount"]
        else:
            expenses_by_category[category] = transaction["amount"]

    return expenses_by_category


try:
    excel = create_excel()
    all_transactions = read_csv_transactions("./CSV/july_22.csv")
    expenses_by_category = group_transactions_by_categories(all_transactions)
    write_expenses_by_category_into_excel(
        excel["sheets"]["summary"], expenses_by_category
    )

    write_transactions_into_excel(excel["sheets"]["spendings"], all_transactions)

    excel["workbook"].save("Spendings.xlsx")
    print("Excel file was successfully saved 🚀")
except Exception as exception:
    print(exception)
    print("Something went wrong while saving excel file 😢")
